"""Normalization layer: openpyxl → InternalGrid.

Responsibilities
----------------
1. Read the workbook via openpyxl (NOT pandas — pandas cannot detect merge cells).
2. Build a merge map so that every cell in a merged region carries the master
   cell's value and is flagged `is_merged=True`.
3. Normalise dates/times to ISO-8601 strings (YYYY-MM-DD / HH:MM).
4. Convert openpyxl's 1-based coordinates to 0-based for the rest of the engine.
"""

from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass
class InternalCell:
    """A single normalised cell.

    value          : str representation of the cell value, or None if empty.
                     Dates are already converted to 'YYYY-MM-DD'.
    original_value : the raw value returned by openpyxl before normalisation.
    is_merged      : True if this cell was expanded from a merge range.
    """
    value: str | None
    original_value: Any
    is_merged: bool = False


# Module-level sentinel used when normalising merged cells that have no
# real openpyxl cell object (avoids creating a class inside a loop).
@dataclass
class _MergedCellProxy:
    """Lightweight stand-in for an openpyxl cell inside a merge range."""
    value: Any
    is_date: bool = False


class InternalGrid:
    """A 2-D array of InternalCell objects with 0-based coordinate access."""

    def __init__(self, cells: list[list[InternalCell]]):
        self._cells = cells
        self.num_rows = len(cells)
        self.num_cols = max((len(row) for row in cells), default=0)

    def get_cell(self, row: int, col: int) -> InternalCell | None:
        """Return the cell at (row, col), or None if out of bounds."""
        if row < 0 or row >= self.num_rows:
            return None
        if col < 0 or col >= len(self._cells[row]):
            return None
        return self._cells[row][col]

    def get_row_slice(self, row: int, start_col: int, length: int) -> list[InternalCell]:
        """Return *length* cells starting at (row, start_col).

        Out-of-bounds positions are filled with empty cells.
        """
        result = []
        for c in range(start_col, start_col + length):
            cell = self.get_cell(row, c)
            if cell is None:
                cell = InternalCell(value=None, original_value=None, is_merged=False)
            result.append(cell)
        return result

    def get_col_slice(self, start_row: int, col: int, length: int) -> list[InternalCell]:
        """Return *length* cells starting at (start_row, col).

        Out-of-bounds positions are filled with empty cells.
        """
        result = []
        for r in range(start_row, start_row + length):
            cell = self.get_cell(r, col)
            if cell is None:
                cell = InternalCell(value=None, original_value=None, is_merged=False)
            result.append(cell)
        return result

    def is_row_all_empty(
        self,
        row: int,
        start_col: int,
        num_cols: int,
        *,
        allow_whitespace: bool = True,
    ) -> bool:
        """Return True if every cell in the given row slice is considered empty.

        Parameters
        ----------
        allow_whitespace : if True (default), cells whose value is an empty or
            whitespace-only string also count as empty (consistent with
            EmptyRow(allow_whitespace=True)).  If False, only None (truly
            absent) cells are considered empty.
        """
        for c in range(start_col, start_col + num_cols):
            cell = self.get_cell(row, c)
            if cell is None:
                continue
            if allow_whitespace:
                if cell.value is not None and cell.value.strip() != "":
                    return False
            else:
                if cell.value is not None:
                    return False
        return True


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------

def _normalise_value(cell, is_date: bool) -> tuple[str | None, Any]:
    """Convert an openpyxl cell value to a normalised string.

    Returns (normalised_str_or_None, original_value).
    """
    import datetime

    original = cell.value
    if original is None:
        return None, None

    if is_date or isinstance(original, (datetime.date, datetime.datetime)):
        if isinstance(original, datetime.datetime):
            return original.strftime("%Y-%m-%d"), original
        if isinstance(original, datetime.date):
            return original.strftime("%Y-%m-%d"), original
        # numeric date serial from openpyxl
        try:
            from openpyxl.utils.datetime import from_excel
            dt = from_excel(original)
            return dt.strftime("%Y-%m-%d"), original
        except Exception:
            pass

    if isinstance(original, datetime.time):
        return original.strftime("%H:%M"), original

    if isinstance(original, float) and original == int(original):
        # e.g. 1000.0 → "1000" so that Types.INT pattern matches
        return str(int(original)), original

    return str(original), original


def load_and_normalize_excel(
    file_path: str | Path,
    sheet: str | int = 0,
) -> tuple[InternalGrid, str]:
    """Load an Excel file and return an InternalGrid plus the resolved sheet name.

    Parameters
    ----------
    file_path : path to the .xlsx / .xls file
    sheet     : sheet name (str) or 0-based sheet index (int)

    Returns
    -------
    (grid, sheet_name)
    """
    path_str = str(file_path)
    if path_str.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path_str, formatting_info=True)
        return _load_xls_from_wb(wb, sheet)
    
    import openpyxl
    wb = openpyxl.load_workbook(path_str, data_only=True)
    res = _load_xlsx_from_wb(wb, sheet)
    wb.close()
    return res

def _load_xlsx_from_wb(
    wb: Any,
    sheet: str | int = 0,
) -> tuple[InternalGrid, str]:
    from openpyxl.cell.cell import MergedCell

    # Resolve sheet
    if isinstance(sheet, int):
        sheet_name = wb.sheetnames[sheet]
    else:
        sheet_name = sheet
    ws = wb[sheet_name]

    # Build merge map: (1-based row, 1-based col) → master value
    merge_map: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        master_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        master_val = master_cell.value
        for row_cells in merged_range.rows:
            for coord in row_cells:
                merge_map[(coord.row, coord.column)] = master_val

    # Build the grid (convert to 0-based)
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    grid_cells: list[list[InternalCell]] = []
    for r1 in range(1, max_row + 1):
        row_data: list[InternalCell] = []
        for c1 in range(1, max_col + 1):
            raw_cell = ws.cell(r1, c1)
            is_merged_cell = isinstance(raw_cell, MergedCell)

            if is_merged_cell:
                master_val = merge_map.get((r1, c1))
                proxy = _MergedCellProxy(value=master_val)
                norm_val, orig_val = _normalise_value(proxy, False)
                internal = InternalCell(value=norm_val, original_value=orig_val, is_merged=True)
            else:
                try:
                    is_date = raw_cell.is_date
                except Exception:
                    is_date = False
                norm_val, orig_val = _normalise_value(raw_cell, is_date)
                # Distinguish None (empty) from "" (space)
                if orig_val == "" or (isinstance(orig_val, str) and orig_val.strip() == "" and orig_val != ""):
                    norm_val = orig_val  # keep empty string as-is
                internal = InternalCell(value=norm_val, original_value=orig_val, is_merged=False)

            row_data.append(internal)
        grid_cells.append(row_data)

    return InternalGrid(grid_cells), sheet_name

def _load_xls_from_wb(
    wb: Any,
    sheet: str | int = 0,
) -> tuple[InternalGrid, str]:
    import xlrd

    if isinstance(sheet, int):
        sh = wb.sheet_by_index(sheet)
        sheet_name = sh.name
    else:
        sh = wb.sheet_by_name(sheet)
        sheet_name = sheet

    # Build merge map: (row, col) 0-based -> master (row, col)
    # xlrd returns merged_cells as list of (row_low, row_high, col_low, col_high)
    merge_map = {}
    master_cells = {}

    for crange in sh.merged_cells:
        rlo, rhi, clo, chi = crange
        # master is a top-left cell
        master_coord = (rlo, clo)
        for rowx in range(rlo, rhi):
            for colx in range(clo, chi):
                merge_map[(rowx, colx)] = master_coord

    max_row = sh.nrows
    max_col = sh.ncols

    grid_cells: list[list[InternalCell]] = []
    import datetime

    for r in range(max_row):
        row_data: list[InternalCell] = []
        for c in range(max_col):
            is_merged_cell = (r, c) in merge_map and merge_map[(r, c)] != (r, c)
            
            if is_merged_cell:
                mr, mc = merge_map[(r, c)]
                raw_val = sh.cell_value(rowx=mr, colx=mc)
                raw_type = sh.cell_type(rowx=mr, colx=mc)
            else:
                raw_val = sh.cell_value(rowx=r, colx=c)
                raw_type = sh.cell_type(rowx=r, colx=c)

            is_date = (raw_type == xlrd.XL_CELL_DATE)
            if raw_val == "":
                raw_val = None

            if raw_val is not None and is_date:
                try:
                    dt_tuple = xlrd.xldate_as_tuple(raw_val, wb.datemode)
                    raw_val = datetime.datetime(*dt_tuple)
                except Exception:
                    pass
            
            orig_val = raw_val
            
            proxy = _MergedCellProxy(value=raw_val)
            norm_val, _ = _normalise_value(proxy, is_date)

            if orig_val == "" or (isinstance(orig_val, str) and orig_val.strip() == "" and orig_val != ""):
                norm_val = orig_val

            internal = InternalCell(value=norm_val, original_value=orig_val, is_merged=is_merged_cell)
            row_data.append(internal)
        grid_cells.append(row_data)

    return InternalGrid(grid_cells), sheet_name
